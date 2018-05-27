from openpyxl import load_workbook
from openpyxl import Workbook
import re
import math

key_pat = '^B([2-9]|[1-9]\d+)'
inwb = load_workbook('original_price.xlsx')
outwb = Workbook()

sheet1 = inwb[inwb.sheetnames[0]]

ws1 = outwb.active
ws1.title = inwb.sheetnames[0]

for row in sheet1.iter_rows():
    ws1['D1'] = '自动算出的价格' 
    for cell in row:
        ws1[cell.coordinate] = cell.value
        result = re.findall(key_pat, cell.coordinate)
        if len(result) > 0:
            sell_price = cell.value * 2
            if (math.floor(sell_price) == round(sell_price)) and (sell_price != round(sell_price)):
                sell_price = math.floor(sell_price) + 0.5
            else:
                sell_price = round(sell_price)

            ws1['D'+ result[0]] = sell_price
    
outwb.save('original_price2.xlsx')