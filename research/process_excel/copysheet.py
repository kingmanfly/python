from openpyxl import load_workbook
from openpyxl import Workbook

inwb = load_workbook('original_price.xlsx')
outwb = Workbook()

insheetnames = inwb.get_sheet_names()#获取sheet页
sheet1 = inwb.get_sheet_by_name(insheetnames[0])

ws1 = outwb.active
ws1.title = insheetnames[0]

for row in sheet1.iter_rows():  
    for cell in row:
        print(cell.coordinate, cell.value, end=',') 
        ws1[cell.coordinate] = cell.value

outwb.save('original_price2.xlsx')



