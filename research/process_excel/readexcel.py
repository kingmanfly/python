from openpyxl import load_workbook

wb = load_workbook('original_price.xlsx')
print(wb.sheetnames)

sheet = wb['Sheet1']  
for row in sheet.iter_rows():  
    for cell in row:  
        print(cell.coordinate, cell.value, end=",")
    print("")
print("~~~~~~~~~~~~~~~~~~~~~~~~")
sheet = wb['Sheet1']
for rowOfCellObjects in sheet['A1':'C3']:  
    for cellObj in rowOfCellObjects:  
        print(cellObj.coordinate, cellObj.value, end=",")
    print("")
        
print('--- END OF ROW ---')
